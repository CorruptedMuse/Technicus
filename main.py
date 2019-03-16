import discord
from discord.ext import commands
import authDeets
import musicBot
import roleBot
import subBot
import modBot
import miscBot
import flagdata
from babel import languages
from googletrans import Translator
import re
from commonFunctions import log
from openpyxl import Workbook
import time
import asyncio
import sqlite3

description = '''The Technicus Bot for the Door Monster Official Server'''
bot = commands.Bot(command_prefix='.', description=description, case_insensitive=True)
version = "version=2.0"
is_voting = False

bot.add_cog(musicBot.Music(bot))
bot.add_cog(roleBot.Role(bot))
bot.add_cog(miscBot.Misc(bot))
bot.add_cog(subBot.Subber(bot))
bot.add_cog(modBot.Mod(bot))
bot.remove_command("help")

client = discord.Client()

connection = sqlite3.connect("reminders.db")
cursor = connection.cursor()

translator = Translator()

translated_messages = []
noodle_list = []

RE_EMOJI = re.compile('[\U00010000-\U0010ffff]', flags=re.UNICODE)

def strip_emoji(text):
    return RE_EMOJI.sub(r'', text)

@bot.event
async def on_ready():
    print("Logged in as: {0}, with the ID of: {1}".format(bot.user, bot.user.id))
    await bot.change_presence(activity=discord.Game(name='with your code'))
    print("--")


@bot.event
async def on_member_join(member):
    await asyncio.sleep(2)
    for role in member.roles:
        if role.name == "linked":
            cursor.execute("SELECT * FROM silenced")
            result = cursor.fetchall()
            is_silenced = False
            for r in result:
                if member.id == r[0]:
                    is_silenced = True
            if is_silenced:
                await member.add_roles(discord.utils.get(member.guild.roles, name="Silenced"))
            else:
                await member.add_roles(discord.utils.get(member.guild.roles, name="Patron"))


@bot.event
async def on_member_update(old_member, new_member):
    if old_member.roles != new_member.roles:
        if set(old_member.roles) - set(new_member.roles) == set():
            role = set(new_member.roles) - set(old_member.roles)
            if str(role.pop()) == "Controversial":
                await old_member.send("""#civil-discussion is a place for those two things exactly--discussion and civility. It's a place to exchange ideas with others and broaden your personal horizons. In accordance with the possibly sensitive nature of some of the discussions on this channel, we have a few extra rules and reminders in place in addition to our general rules, which are still very much in effect, as well.

0. This channel has a Two Strike policy. Before anything else, be aware that this channel in particular has a different standard of moderation and consequence. If you commit an infraction of these rules, you will receive one, and only one, warning; upon your second infraction, your permission to participate in this channel will be revoked. If you have any uncertainty about whether or not what you have to say will cross the line, play it safe, scale your comment back, and temper it with kindness.

1. Keep your conversations about ideas, not about people. Think of this as an elaboration on our "no personal attacks" rule. There will be zero tolerance of attacks or criticisms made against specific people (or indeed, groups of specific people) on the basis of their lifestyle, gender, race, religion, culture, or identity. If you want to discuss one of these aspects of someone, you must do it as an address of the idea at large and leave the individual out of it.""")
                await old_member.send("""2. Discussion is to be encouraged and facilitated, not circumvented or silenced. This is a place for conversations, and more importantly, conversations carried out in good faith. It is necessary, therefore, that all posts be made with the goal and understanding of receiving a response--and quite likely one with a different point of view or opinion. When you go to make a post, think first to yourself, "Am I interested in what other people have to say about this?" No one should operate here under the impression that they will be able to have the "last word" on a given subject, and anyone who tries to prevent or discourage another from expressing their input will receive a strike.

3. Remember that participation is voluntary. This is an opt-in channel, and no one is required to participate in or comment on anything that they don't feel they would like to. In any discussion, at any point, you always have the option to step back and disengage. Have a disagreement with someone and neither of you seem like you're going to budge? Step back and agree to disagree once the conversation has run its course. Someone having a conversation that makes you uncomfortable or angry? Disengage and come back later. All posters are expected to handle themselves with responsibility and maturity, and that means knowing when not to post as much as it does knowing what and how to post.""")
                await old_member.send("""4. Act always with kindness. Lastly, simply remember that we are all human, and we all deserve kindness. Consider the viewpoints of others with compassion and empathy, and respond to everyone, regardless of agreement or stance, with the goal of improving the discourse. These discussions are supposed to be interesting, they're supposed to be enlightening and comfortable, and hey, they're supposed to be fun.""")
            log("Added Role", old_member, None, set(new_member.roles) - set(old_member.roles), None)
        else:
            log("Deleted Role", old_member, None, set(old_member.roles) - set(new_member.roles), None)
    old_not_patreon = True
    old_not_kick = True
    new_is_patreon = False
    new_is_kick = False
    is_roled = False
    was_not_roled = True
    for role in old_member.roles:
        if role.name == "linked":
            old_not_patreon = False
        if role.name == "Patron":
            was_not_roled = False
        if role.name == "KickstarterBacker":
            old_not_kick = False
    for role in new_member.roles:
        if role.name == "KickstarterBacker":
            new_is_kick = True
        if role.name == "Patron":
            is_roled = True
        if role.name == "linked":
            new_is_patreon = True
    if old_not_patreon and new_is_patreon and not is_roled:
        cursor.execute("SELECT * FROM silenced")
        result = cursor.fetchall()
        is_silenced = False
        for r in result:
            if new_member.id == r[0]:
                is_silenced = True
        if is_silenced:
            await new_member.add_roles(discord.utils.get(new_member.guild.roles, name="Silenced"))
        else:
            await new_member.add_roles(discord.utils.get(new_member.guild.roles, name="Patron"))
    if was_not_roled and old_not_kick and (is_roled or new_is_kick):
        await welcome(old_member)


async def welcome(member):
    guild = member.guild
    join_embed = discord.Embed(title="{} has joined the server.".format(member),
                              description='Join Date: {} UTC'.format(str(member.joined_at)[:19]), color=discord.Color.green())
    join_embed.set_footer(text='User Joined')
    join_embed.set_thumbnail(url=member.avatar_url)
    await discord.utils.get(guild.channels, name='introductions').send(embed=join_embed)
    log("Joined", member, None, None, None)


@bot.event
async def on_member_remove(member):
    is_roled = False
    for role in member.roles:
        if role.name == "Patron" or role.name == "KickstarterBacker":
            is_roled = True
    if is_roled:
        guild = member.guild
        cursor.execute("DELETE FROM reminders WHERE authorID={0}".format(member.id))
        cursor.execute("DELETE FROM userInfo WHERE authorID={0}".format(member.id))
        connection.commit()
        leave_embed = discord.Embed(title="{} has left the server.".format(member),
                                   description='Leave Date: {} UTC'.format(time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime())),
                                   color=discord.Color.red())
        leave_embed.set_footer(text='User Left')
        leave_embed.set_thumbnail(url=member.avatar_url)
        await discord.utils.get(member.guild.channels, name='introductions').send(embed=leave_embed)
        log("Left Guild", member, None, None, None)


@bot.event
async def on_message_delete(message):
    log("Message Delete", message.author, message.channel, message.content, message.created_at)


@bot.event
async def on_message_edit(message1, message2):
    if message1.content == message2.content:
        return
    log("Edited Message", message1.author, message1.channel, message1.content, message2.content)
    
@bot.event
async def on_raw_reaction_add(payload):
    emoji = payload.emoji
    message_id = payload.message_id
    channel_id = payload.channel_id
    member_id = payload.user_id
    channel = bot.get_channel(channel_id)
    
    the_message = None
    async for message in channel.history(limit=50):
        if message.id == message_id:
            the_message = message
    if the_message:
        member = the_message.guild.get_member(member_id)
        member_permissions = channel.permissions_for(member)
        
        cursor.execute("SELECT * FROM botBans")
        result = cursor.fetchall()
        for r in result:
            if member.id == r[0]:
                return
        
        if member_permissions.send_messages and emoji.is_unicode_emoji() and the_message.content is not "":
            for flag in flagdata.flags:
                if emoji.name == flag['emoji']:
                    if flag['code'] == "US":
                        lang_code = "en"
                    else:
                        try:
                            lang_code = languages.get_official_languages(flag['code'])[0]
                        except:
                            return
                    if lang_code:
                        list_pos = 1
                        has_entry = False
                        while list_pos < len(translated_messages) and not has_entry:
                            has_entry = translated_messages[list_pos][0] == message_id
                            list_pos = list_pos + 1
                        list_pos = list_pos - 1
                        if has_entry:
                            if lang_code in translated_messages[list_pos][1]:
                                return
                            translated_messages[list_pos][1].append(lang_code)
                        else:
                            translated_messages.append([message_id, [lang_code]])
                        translated = translator.translate(strip_emoji(the_message.content), dest=lang_code)
                        if translated.src == lang_code:
                            return
                            
                        is_mod = False
                        for role in member.roles:
                            if role.name == "Bot Mod":
                                is_mod = True
                        
                        trans_embed = discord.Embed(title="Translating from **{0}** to **{1}**...".format(translated.src, lang_code), description=translated.text)
                        trans_embed.set_author(name=the_message.author.display_name, icon_url=the_message.author.avatar_url)
                        trans_embed.set_footer(text="Requested by {}".format(member.display_name), icon_url=member.avatar_url)
                        if channel.name != "bot-commands" and not is_mod:
                            bot_silenced = add_spam_time(member.id)
                            if bot_silenced:
                                await member.send("Do not spam the bot outside #bot-commands, please wait a bit before using the bot again")
                            else:
                                await channel.send(embed=trans_embed)
                        else:
                            await channel.send(embed=trans_embed)

@bot.command()
async def help(ctx, *args_command_name):
    if len(args_command_name) == 0:
        msg = "```This is the Technicus Bot for the Door Monster server! Type .help <command> to get information on the following commands:\n\n"
        for command in bot.all_commands:
            if command not in ["musicbot", "editrole", "bridge", "massdel", "botban", "silence", "lock"]:
                msg = "{0}{1}\n".format(msg, command)
        msg = "{0}\n Contact CorruptedMuse for additional support```".format(msg)
        return await ctx.send(msg)
    else:
        command_name = args_command_name[0]
        command_docs = open("commandDocumentation.txt", "r")
        msg = "```"
        rec = False
        for command_docs_line in command_docs:
            if not rec and command_name == command_docs_line[:-1]:
                msg = "{0}{1}\n".format(msg, command_docs_line[:-1])
                rec = True
            if rec and command_docs_line[:-1] == "-":
                rec = False
            if rec:
                msg = "{0}{1}\n".format(msg, command_docs_line[:-1])
        msg = "{}```".format(msg)
        if msg == "``````":
            return await ctx.send("**Error:** Command not found")
        await ctx.send(msg)

@bot.event
async def on_message(message):
    cursor.execute("SELECT * FROM botBans")
    result = cursor.fetchall()
    is_bot_banned = False
    for r in result:
        if message.author.id == r[0]:
            is_bot_banned = True
    if message.author.id != bot.user.id and not is_bot_banned:
        is_mod = False
        for role in message.author.roles:
            if role.name == "Bot Mod":
                is_mod = True

        if message.author.id == 227187657715875841 and message.content.lower()=="system;stats":
            wb = Workbook()
            ws = wb.active
            row = 1
            await message.channel.send("Storing user info...")
            for member in message.guild.members:
                ws['A{}'.format(row)] = member.joined_at
                ws['B{}'.format(row)] = member.id
                ws['C{}'.format(row)] = member.name
                row = row + 1
            ws1 = wb.create_sheet("Channels")
            row = 1
            #for channel in message.guild.text_channels:
            #    await message.channel.send("""Storing channel <#{}>...""".format(channel.id))
            #    try:
            #        async for a_message in channel.history(limit=1000000):
            #            ws1['A{}'.format(row)] = a_message.created_at
            #            ws1['B{}'.format(row)] = a_message.channel.id
            #            ws1['C{}'.format(row)] = a_message.channel.name
            #            ws1['D{}'.format(row)] = a_message.id
            #            ws1['E{}'.format(row)] = a_message.author.id
            #            ws1['F{}'.format(row)] = a_message.author.name
            #            ws1['G{}'.format(row)] = a_message.content
            #            row=row+1
            #    except:
            #        await message.channel.send("Error in storing channel")
            wb.save("data1.xlsx")
            await message.channel.send("Complete.")
                
        if message.channel.name != "bot-commands" and message.content.startswith('.') and not is_mod:
            silenced = add_spam_time(message.author.id)
            if silenced:
                await message.author.send(
                    "Do not spam the bot outside #bot-commands, please wait a bit before using the bot again")
            else:
                await bot.process_commands(message)
        else:
            await bot.process_commands(message)


async def remind_user():
    await bot.wait_until_ready()
    while not bot.is_closed():
        to_be_deleted = []
        cursor.execute("SELECT * FROM reminders")
        result = cursor.fetchall()
        for r in result:
            if r[4] < time.time():
                the_channel = bot.get_channel(r[1])
                if the_channel is not None:
                    await the_channel.send("<@{0}> {1}".format(r[2], r[3].replace("Ё", "'")))
                else:
                    log("Reminder Sys Error", None, None, r, None)
                to_be_deleted.append(r[0])
        for reminderID in to_be_deleted:
            cursor.execute("DELETE FROM reminders WHERE reminderID={0}".format(reminderID))
            connection.commit()
        await asyncio.sleep(.5)


async def spamreduce():
    await bot.wait_until_ready()
    while not bot.is_closed():
        list_pos = 0
        to_be_deleted = []
        while list_pos < len(noodle_list):
            noodle_list[list_pos][1] = noodle_list[list_pos][1] - 5
            if noodle_list[list_pos][1] < 0:
                to_be_deleted.append(list_pos)
            list_pos = list_pos + 1
        for noodlePos in reversed(to_be_deleted):
            del noodle_list[noodlePos]
        await asyncio.sleep(5)


def add_spam_time(author_id):
    list_pos = 0
    has_entry = False
    while list_pos < len(noodle_list) and not has_entry:
        has_entry = noodle_list[list_pos][0] == author_id
        list_pos = list_pos + 1
    list_pos = list_pos - 1
    if has_entry:
        if noodle_list[list_pos][1] > 200:
            return True
        noodle_list[list_pos][1] = noodle_list[list_pos][1] + 60
        return False
    else:
        noodle_list.append([author_id, 60])
        return False

bot.loop.create_task(remind_user())
bot.loop.create_task(spamreduce())
bot.run(authDeets.token)
