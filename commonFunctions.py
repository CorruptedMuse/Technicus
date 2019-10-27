from openpyxl import load_workbook, Workbook
import time
import discord

def merge_strings(string_array):
    result = ""
    for segment in string_array:
        result = "{0}{1} ".format(result, segment)
    return result[:-1]


def log(event, user, channel, state1, state2):
    # Log format: time, event, user, channel, first state, second state
    print("Recording a {}".format(event))
    monthyear = time.strftime("%b%y", time.gmtime())
    try:
        wb = load_workbook("logs{}.xlsx".format(monthyear))
    except:
        wb = Workbook()
    sheet_name = monthyear
    ws = None
    for sheet in wb:
        if sheet.title == sheet_name:
            ws = sheet
    if not ws:
        ws = wb.create_sheet(sheet_name)

    row = ws['I1'].value
    if row is None:
        row = 1
    while ws['A{}'.format(row)].value is not None:
        row = row + 1
    ws['I1'].value = row
    ws['A{}'.format(row)] = time.asctime(time.gmtime())
    ws['B{}'.format(row)] = event
    if user is not None:
        ws['C{}'.format(row)] = user.name
        ws['D{}'.format(row)] = user.id
    if channel is not None:
        ws['E{}'.format(row)] = channel.name
        ws['F{}'.format(row)] = channel.id
    ws['G{}'.format(row)] = str(state1)
    ws['H{}'.format(row)] = str(state2)
    wb.save("logs{}.xlsx".format(monthyear))

def find_user(message, args):
    member = discord.utils.get(message.guild.members, name=merge_strings(args))
    if member is None:
        member = discord.utils.get(message.guild.members, nick=merge_strings(args))
    if member is None and len(message.mentions) is not 0:
        member = message.mentions[0]
    return member